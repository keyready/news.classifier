import json
from openpyxl import load_workbook

def extract_data_from_xlsx(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    texts = []
    categories = []

    for row in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 пропускает заголовки
        texts.append(row[0])
        categories.append(row[1])

    data = {
        'text': texts,
        'category': categories
    }

    return data

def save_data_to_json(data, output_file):
    # Сохраняем данные в JSON-файл
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

# Пример использования функции
file_path = 'news-lenta-cleaned-all.xlsx'  # Замените на путь к вашему файлу
output_file = 'output_data.json'  # Путь для сохранения результата

# Извлекаем данные
data = extract_data_from_xlsx(file_path)

# Сохраняем данные в файл
if data:
    save_data_to_json(data, output_file)
    print(f"Данные успешно сохранены в {output_file}")
